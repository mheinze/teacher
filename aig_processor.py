#!/usr/bin/env python3
"""
AIG Class List Processor

This application processes class lists from PDF files and Word documents, then combines them with Excel spreadsheets
to generate classroom-specific reports showing AIG (Academically/Intellectually Gifted) status.

Features:
- Extracts student data from PDF files with columns: Name, Student Id, Grade, Reading, Math
- Extracts student data from Word documents with table format: Name, Reading, Math
- Combines data with Excel spreadsheets (uses only PDF/Word data for AIG status)
- Handles grade, track, and classroom information from Excel first rows
- Updates original Excel file by adding AIG columns to each sheet
- Color codes rows: Blue for AIG Math only, Orange for AIG Reading only, Yellow for both
- Treats TD, AG, IG and AIG all as AIG status in both PDF and Word data
- Removes all temporary files, leaving only the final Excel spreadsheet
"""

import pandas as pd
import PyPDF2
import os
import re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import logging
from docx import Document

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class AIGClassListProcessor:
    def __init__(self, pdf_file_path, excel_file_path, word_file_path=None, output_dir='output'):
        """
        Initialize the processor with file paths
        Updated to save files to output directory
        
        Args:
            pdf_file_path (str): Path to the PDF file containing AIG roster
            excel_file_path (str): Path to the Excel file containing class lists
            word_file_path (str): Optional path to Word document with additional AIG students
            output_dir (str): Directory to save output files
        """
        self.pdf_file_path = pdf_file_path
        self.excel_file_path = excel_file_path
        self.word_file_path = word_file_path
        self.output_dir = output_dir
        
        # Create output directory if it doesn't exist
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)
            
        # Set output file paths
        self.output_file = os.path.join(self.output_dir, 'updated_class_lists.xlsx')
        self.aig_only_file = os.path.join(self.output_dir, 'updated_class_lists_AIG_Only.xlsx')
        self.missing_students_file = os.path.join(self.output_dir, 'students_not_in_excel.xlsx')
        
        self.aig_students = {'math': set(), 'reading': set()}
        self.class_lists = {}
        
        # Track students found in Excel vs source documents
        self.students_in_excel = set()
        self.students_in_sources = set()  # From PDF and Word documents
        
        # Statistics tracking
        self.stats = {
            'total_students': 0,
            'aig_math_only': 0,
            'aig_reading_only': 0,
            'aig_both': 0,
            'aig_none': 0,
            'td_only_students': set()  # Track students who are only TD (not AG, IG, AIG)
        }
        
        # Define color fills - Updated per new requirements
        self.colors = {
            'math': PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid'),      # Light blue for math (as specified)
            'reading': PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid'),  # Orange for reading
            'both': PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')      # Yellow for both
        }
    
    def extract_aig_students_from_pdf(self):
        """
        Extract AIG student names from the PDF file
        """
        logger.info(f"Processing PDF file: {self.pdf_file_path}")
        
        try:
            with open(self.pdf_file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                text = ""
                for page in pdf_reader.pages:
                    text += page.extract_text()
                
                # Parse the text to extract student names and their AIG subjects
                self._parse_aig_text(text)
                
        except Exception as e:
            logger.error(f"Error reading PDF file: {e}")
            raise
    
    def _parse_aig_text(self, text):
        """
        Parse the extracted text to identify AIG students and their subjects
        PDF format: Name Student Id Grade Reading Math
        Where Reading/Math can be: AIG, AG, IG, TD, or -
        Note: TD, AG, IG and AIG all are treated as AIG
        
        Args:
            text (str): Raw text extracted from PDF
        """
        lines = text.split('\n')
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
            
            # Skip header lines
            if 'Name Student Id Grade Reading Math' in line or 'School Roster' in line:
                continue
                
            # Try to parse student data line
            # Expected format: "Last, First StudentId Grade Reading Math"
            if ',' in line:  # Student names have commas
                parts = line.split()
                
                # Find the comma in the name to split name from other data
                name_parts = []
                data_parts = []
                comma_found = False
                
                for part in parts:
                    if not comma_found:
                        name_parts.append(part)
                        if ',' in part:
                            comma_found = True
                    else:
                        data_parts.append(part)
                
                # We should have: Name parts + [FirstName, StudentId, Grade, Reading, Math]
                # But some students have compound first names, so we need better parsing
                
                # Find the student ID (long numeric string) to anchor the parsing
                student_id_idx = -1
                for i, part in enumerate(data_parts):
                    if part.isdigit() and len(part) >= 8:  # Student IDs are long
                        student_id_idx = i
                        break
                
                if student_id_idx >= 1:  # At least one name part before student ID
                    try:
                        # Everything before student ID is first name
                        first_name_parts = data_parts[:student_id_idx]
                        first_name = ' '.join(first_name_parts)
                        
                        # After student ID: grade, reading, math
                        remaining_parts = data_parts[student_id_idx:]
                        if len(remaining_parts) >= 4:  # ID, Grade, Reading, Math
                            student_id = remaining_parts[0]
                            grade = remaining_parts[1]
                            reading_status = remaining_parts[2]
                            math_status = remaining_parts[3]
                            
                            # Reconstruct full name
                            full_name = ' '.join(name_parts) + ' ' + first_name
                            normalized_name = self._normalize_name(full_name)
                            
                            # Parse AIG status - Updated: TD, AG, IG and AIG all are the same as AIG
                            is_aig_reading = reading_status.upper() in ['TD', 'AG', 'IG', 'AIG']
                            is_aig_math = math_status.upper() in ['TD', 'AG', 'IG', 'AIG']
                            
                            # Track TD-only students (only TD, not AG/IG/AIG)
                            is_td_only_reading = reading_status.upper() == 'TD'
                            is_td_only_math = math_status.upper() == 'TD'
                            if (is_td_only_reading or is_td_only_math) and \
                               not (reading_status.upper() in ['AG', 'IG', 'AIG'] or math_status.upper() in ['AG', 'IG', 'AIG']):
                                self.stats['td_only_students'].add(normalized_name)
                            
                            # Store student details
                            student_info = {
                                'name': normalized_name,
                                'student_id': student_id,
                                'grade': grade,
                                'reading': is_aig_reading,
                                'math': is_aig_math
                            }
                            
                            # Add to AIG sets
                            if is_aig_math:
                                self.aig_students['math'].add(normalized_name)
                            if is_aig_reading:
                                self.aig_students['reading'].add(normalized_name)
                                
                            # Store detailed student info
                            if not hasattr(self, 'student_details'):
                                self.student_details = {}
                            self.student_details[normalized_name] = student_info
                            
                            # Track student from source document
                            self.students_in_sources.add(normalized_name)
                            
                            logger.debug(f"Added student: {normalized_name}, Grade: {grade}, Math: {is_aig_math}, Reading: {is_aig_reading}")
                            
                    except (ValueError, IndexError) as e:
                        logger.debug(f"Could not parse line: {line} - {e}")
                        continue
    
    def _is_student_name(self, line):
        """
        Determine if a line contains a student name
        
        Args:
            line (str): Line of text to check
            
        Returns:
            bool: True if line appears to contain a student name
        """
        # Simple heuristic: contains comma and alphabetic characters
        return ',' in line and any(c.isalpha() for c in line) and len(line.split()) >= 2
    
    def _normalize_name(self, name):
        """
        Normalize student name for consistent matching
        
        Args:
            name (str): Raw name string
            
        Returns:
            str: Normalized name
        """
        # Remove extra whitespace and convert to title case
        name = re.sub(r'\s+', ' ', name.strip())
        return name.title()
    
    def extract_aig_students_from_word(self):
        """
        Extract AIG student names from the Word document
        Word format: Table with three columns - Name, Reading, Math
        First column: Name (comma format "last, first" or "first last")
        Other columns: Reading and Math AIG status (TD means AIG)
        """
        if not self.word_file_path or not os.path.exists(self.word_file_path):
            logger.info("No Word document provided or file not found, skipping Word processing")
            return
            
        logger.info(f"Processing Word document: {self.word_file_path}")
        
        try:
            doc = Document(self.word_file_path)
            
            # Process all tables in the document
            for table in doc.tables:
                # Skip header row
                for i, row in enumerate(table.rows):
                    if i == 0:  # Skip header row
                        continue
                        
                    cells = row.cells
                    if len(cells) >= 3:
                        name_text = cells[0].text.strip()
                        reading_text = cells[1].text.strip()
                        math_text = cells[2].text.strip()
                        
                        if name_text:
                            # Process name format
                            normalized_name = self._process_word_name(name_text)
                            
                            # Check AIG status (TD means AIG in Word document)
                            is_aig_reading = reading_text.upper() in ['TD', 'AIG']
                            is_aig_math = math_text.upper() in ['TD', 'AIG']
                            
                            # Track TD-only students from Word document
                            is_td_only_reading = reading_text.upper() == 'TD'
                            is_td_only_math = math_text.upper() == 'TD'
                            if (is_td_only_reading or is_td_only_math) and reading_text.upper() != 'AIG' and math_text.upper() != 'AIG':
                                self.stats['td_only_students'].add(normalized_name)
                            
                            # Store student details
                            if not hasattr(self, 'student_details'):
                                self.student_details = {}
                                
                            # Update or create student record
                            if normalized_name in self.student_details:
                                # Merge with existing data from PDF
                                self.student_details[normalized_name]['reading'] = self.student_details[normalized_name]['reading'] or is_aig_reading
                                self.student_details[normalized_name]['math'] = self.student_details[normalized_name]['math'] or is_aig_math
                            else:
                                # Create new record
                                self.student_details[normalized_name] = {
                                    'name': normalized_name,
                                    'student_id': 'N/A',  # Not available in Word doc
                                    'grade': 'N/A',       # Not available in Word doc
                                    'reading': is_aig_reading,
                                    'math': is_aig_math
                                }
                            
                            # Add to AIG sets
                            if is_aig_math:
                                self.aig_students['math'].add(normalized_name)
                            if is_aig_reading:
                                self.aig_students['reading'].add(normalized_name)
                            
                            # Track student from source document
                            self.students_in_sources.add(normalized_name)
                                
                            logger.debug(f"Added from Word: {normalized_name}, Math: {is_aig_math}, Reading: {is_aig_reading}")
                        
        except Exception as e:
            logger.error(f"Error reading Word document: {e}")
            # Don't raise - continue with PDF data only
    
    def _process_word_name(self, name_text):
        """
        Process name from Word document according to the format rules
        
        Args:
            name_text (str): Raw name from Word document
            
        Returns:
            str: Normalized name in "Last, First" format
        """
        name_text = name_text.strip()
        
        if ',' in name_text:
            # Already in "last, first" format
            return self._normalize_name(name_text)
        else:
            # "first last" format - need to convert to "last, first"
            parts = name_text.split()
            if len(parts) >= 2:
                first = parts[0]
                last = ' '.join(parts[1:])  # Handle multiple last name parts
                return self._normalize_name(f"{last}, {first}")
            else:
                # Single name - return as is
                return self._normalize_name(name_text)
    
    def _find_student_in_aig(self, student_name, classroom_grade=None):
        """
        Check if a student is in AIG programs using only PDF data
        Now includes grade matching for better accuracy
        
        Args:
            student_name (str): Student name to check
            classroom_grade (str): Grade of the classroom for matching
            
        Returns:
            dict: Dictionary with 'math' and 'reading' boolean values
        """
        normalized_name = self._normalize_name(str(student_name))
        
        # Initialize results
        result = {'math': False, 'reading': False}
        
        # If we have detailed student info from PDF, use it
        if hasattr(self, 'student_details') and normalized_name in self.student_details:
            student_info = self.student_details[normalized_name]
            
            # Check grade match if provided
            if classroom_grade and student_info['grade'] != str(classroom_grade):
                logger.debug(f"Grade mismatch for {normalized_name}: PDF grade {student_info['grade']} vs classroom grade {classroom_grade}")
                # Still return the AIG status but log the mismatch
            
            result['math'] = student_info['math']
            result['reading'] = student_info['reading']
            return result
        
        # Fallback to name matching if detailed info not available
        # Try exact match first
        in_math = normalized_name in self.aig_students['math']
        in_reading = normalized_name in self.aig_students['reading']
        
        # If no exact match, try partial matching
        if not in_math:
            for aig_name in self.aig_students['math']:
                if self._names_match(normalized_name, aig_name):
                    in_math = True
                    break
        
        if not in_reading:
            for aig_name in self.aig_students['reading']:
                if self._names_match(normalized_name, aig_name):
                    in_reading = True
                    break
        
        return {'math': in_math, 'reading': in_reading}
    
    def _names_match(self, name1, name2):
        """
        Check if two names are likely the same person
        
        Args:
            name1 (str): First name
            name2 (str): Second name
            
        Returns:
            bool: True if names likely match
        """
        # Simple matching: check if last names and first names have significant overlap
        parts1 = name1.replace(',', '').split()
        parts2 = name2.replace(',', '').split()
        
        if len(parts1) >= 2 and len(parts2) >= 2:
            # Check if last names match (assuming first part is last name)
            if parts1[0].lower() == parts2[0].lower():
                # Check if first names match or are similar
                first1 = parts1[1].lower()
                first2 = parts2[1].lower()
                return first1 == first2 or first1.startswith(first2) or first2.startswith(first1)
        
        return False
    
    def update_excel_with_aig_data(self):
        """
        Update the original Excel file by adding AIG columns to each sheet
        New approach: Modify existing Excel file instead of creating separate reports
        Preserves original sheet colors/formatting
        """
        logger.info("Updating Excel file with AIG data...")
        
        # Read the original Excel file with openpyxl to preserve formatting
        
        try:
            # Load original workbook to preserve formatting and sheet colors
            original_wb = load_workbook(self.excel_file_path)
            original_sheet_colors = {}
            
            # Store original sheet colors
            for sheet_name in original_wb.sheetnames:
                sheet = original_wb[sheet_name]
                if hasattr(sheet.sheet_properties, 'tabColor') and sheet.sheet_properties.tabColor:
                    original_sheet_colors[sheet_name] = sheet.sheet_properties.tabColor
            
            original_wb.close()
            
        except Exception as e:
            logger.warning(f"Could not load original workbook for color preservation: {e}")
            original_sheet_colors = {}
        
        # Read the original Excel file with pandas
        excel_file = pd.ExcelFile(self.excel_file_path)
        
        # Create new workbook for output
        wb = Workbook()
        
        # Create second workbook for AIG-only students
        wb_aig_only = Workbook()
        
        # Remove default sheets
        if wb.active:
            wb.remove(wb.active)
        if wb_aig_only.active:
            wb_aig_only.remove(wb_aig_only.active)
        
        for sheet_name in excel_file.sheet_names:
            # Read raw data without headers
            df = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)
            
            if len(df) < 3:  # Need at least metadata, header, and one student row
                continue
            
            # Extract classroom info from first row
            classroom_info = self._extract_classroom_info(df.iloc[0], sheet_name)
            
            # Create new worksheet
            ws = wb.create_sheet(title=str(sheet_name))
            ws_aig_only = wb_aig_only.create_sheet(title=str(sheet_name))
            
            # Restore original sheet color if it existed
            if sheet_name in original_sheet_colors:
                ws.sheet_properties.tabColor = original_sheet_colors[sheet_name]
                ws_aig_only.sheet_properties.tabColor = original_sheet_colors[sheet_name]
            
            # Add the metadata row (row 0)
            metadata_row = df.iloc[0].tolist()
            ws.append(metadata_row)
            ws_aig_only.append(metadata_row)
            
            # Add headers (row 1) - keep original headers plus add AIG columns
            header_row = ['LASTNAME', 'FIRSTNAME', 'AIG Math', 'AIG Reading', 'AIG Status']
            ws.append(header_row)
            ws_aig_only.append(header_row)
            
            # Process student data (rows 2+)
            student_data = df.iloc[2:, :2].copy()  # Only first two columns
            student_data = student_data.dropna(how='all')  # Remove empty rows
            
            aig_row_idx = 3  # Track row index for AIG-only sheet
            
            # Add AIG data for each student
            for idx, (_, row) in enumerate(student_data.iterrows(), start=3):  # Start at row 3
                lastname = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
                firstname = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ""
                
                if lastname and firstname and lastname.upper() not in ['LASTNAME', 'LAST']:
                    # Track total students
                    self.stats['total_students'] += 1
                    
                    # Construct full name for matching
                    full_name = f"{lastname}, {firstname}"
                    
                    # Track student found in Excel
                    self.students_in_excel.add(self._normalize_name(full_name))
                    
                    # Find AIG status
                    aig_status = self._find_student_in_aig(full_name, classroom_info['grade'])
                    
                    # Track statistics
                    if aig_status['math'] and aig_status['reading']:
                        status_text = 'Both Math & Reading'
                        color = self.colors['both']
                        self.stats['aig_both'] += 1
                    elif aig_status['math']:
                        status_text = 'Math Only'
                        color = self.colors['math']
                        self.stats['aig_math_only'] += 1
                    elif aig_status['reading']:
                        status_text = 'Reading Only'
                        color = self.colors['reading']
                        self.stats['aig_reading_only'] += 1
                    else:
                        status_text = 'None'
                        color = None
                        self.stats['aig_none'] += 1
                    
                    # Add row data to main sheet
                    row_data = [lastname, firstname, aig_status['math'], aig_status['reading'], status_text]
                    ws.append(row_data)
                    
                    # Apply color formatting to main sheet
                    if color:
                        for col_idx in range(1, 6):  # Columns A through E
                            ws.cell(row=idx, column=col_idx).fill = color
                    
                    # Add to AIG-only sheet if student has AIG status
                    if aig_status['math'] or aig_status['reading']:
                        ws_aig_only.append(row_data)
                        
                        # Apply color formatting to AIG-only sheet
                        if color:
                            for col_idx in range(1, 6):  # Columns A through E
                                ws_aig_only.cell(row=aig_row_idx, column=col_idx).fill = color
                        aig_row_idx += 1
            
            # Auto-adjust column widths for both sheets
            for ws_current in [ws, ws_aig_only]:
                for column in ws_current.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws_current.column_dimensions[column_letter].width = adjusted_width
        
        # Save both Excel files to output directory
        wb.save(self.output_file)
        wb_aig_only.save(self.aig_only_file)
        
        logger.info(f"Updated Excel file saved: {self.output_file}")
        logger.info(f"AIG-only Excel file saved: {self.aig_only_file}")
    
    def _extract_classroom_info(self, first_row, sheet_name):
        """
        Extract classroom information from the first row of a sheet
        
        Args:
            first_row: First row data from Excel sheet
            sheet_name: Name of the sheet
            
        Returns:
            dict: Classroom information
        """
        classroom_info = {
            'grade': None,
            'track': None,
            'teacher': None,
            'sheet_name': sheet_name
        }
        
        # Parse metadata from first row
        first_row_text = ' '.join([str(val) for val in first_row.values if pd.notna(val)])
        
        # Extract grade
        grade_match = re.search(r'(\d+)(?:st|nd|rd|th)?\s*grade|grade\s*(\d+)', first_row_text.lower())
        if grade_match:
            classroom_info['grade'] = grade_match.group(1) or grade_match.group(2)
        
        # Extract track
        track_match = re.search(r't(\d+)|track\s*(\d+)', first_row_text.lower())
        if track_match:
            classroom_info['track'] = track_match.group(1) or track_match.group(2)
        
        # Extract teacher name
        teacher_match = re.search(r'([a-zA-Z]+)(?:\s+\d+(?:st|nd|rd|th)?\s*grade)?', first_row_text)
        if teacher_match:
            classroom_info['teacher'] = teacher_match.group(1)
        else:
            classroom_info['teacher'] = sheet_name
        
        return classroom_info
    
    def _construct_full_name_from_excel(self, row, columns):
        """
        Construct full name from Excel row data, handling separate first/last name columns
        Updated for new Excel structure: LASTNAME, FIRSTNAME columns
        
        Args:
            row: DataFrame row containing student data
            columns: DataFrame columns
            
        Returns:
            str: Constructed full name in "Last, First" format to match PDF
        """
        # With the new structure, we have LASTNAME and FIRSTNAME columns
        lastname = None
        firstname = None
        
        # Try to get values by column name
        if 'LASTNAME' in row:
            lastname = str(row['LASTNAME']).strip() if pd.notna(row['LASTNAME']) else None
        if 'FIRSTNAME' in row:
            firstname = str(row['FIRSTNAME']).strip() if pd.notna(row['FIRSTNAME']) else None
        
        # Fallback: use positional access
        if not lastname and len(row) >= 1:
            lastname = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else None
        if not firstname and len(row) >= 2:
            firstname = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else None
        
        # Construct full name if both parts available
        if lastname and firstname and lastname.upper() not in ['LASTNAME', 'LAST'] and firstname.upper() not in ['FIRSTNAME', 'FIRST']:
            return f"{lastname}, {firstname}"
        
        return None
    
    def print_statistics(self):
        """
        Print statistics to standard output and save to markdown file as required by prompt.md
        """
        print("\n" + "="*60)
        print("AIG STUDENT STATISTICS")
        print("="*60)
        
        # Calculate total AIG students
        total_aig = self.stats['aig_math_only'] + self.stats['aig_reading_only'] + self.stats['aig_both']
        
        print(f"Total students processed in Excel: {self.stats['total_students']}")
        print(f"Students with AIG status: {total_aig}")
        print(f"  - Math only: {self.stats['aig_math_only']}")
        print(f"  - Reading only: {self.stats['aig_reading_only']}")
        print(f"  - Both Math & Reading: {self.stats['aig_both']}")
        print(f"Students with no AIG status: {self.stats['aig_none']}")
        print(f"Students who are only TD: {len(self.stats['td_only_students'])}")
        
        # Missing students information
        missing_count = len(self.students_in_sources - self.students_in_excel)
        print(f"\nSource Document Analysis:")
        print(f"Total students found in PDF/Word: {len(self.students_in_sources)}")
        print(f"Students found in Excel: {len(self.students_in_excel)}")
        print(f"Students in source docs but NOT in Excel: {missing_count}")
        
        # Show percentage
        if self.stats['total_students'] > 0:
            aig_percentage = (total_aig / self.stats['total_students']) * 100
            td_percentage = (len(self.stats['td_only_students']) / self.stats['total_students']) * 100
            print(f"\nPercentages:")
            print(f"  - AIG students: {aig_percentage:.1f}% of Excel total")
            print(f"  - TD-only students: {td_percentage:.1f}% of Excel total")
        
        if len(self.students_in_sources) > 0:
            match_percentage = (len(self.students_in_excel) / len(self.students_in_sources)) * 100
            print(f"  - Students found in Excel: {match_percentage:.1f}% of source total")
        
        print("="*60)
        
        # Save statistics to markdown file as required by prompt.md
        self.save_statistics_to_markdown(total_aig, missing_count, aig_percentage if self.stats['total_students'] > 0 else 0, 
                                       td_percentage if self.stats['total_students'] > 0 else 0, 
                                       match_percentage if len(self.students_in_sources) > 0 else 0)

    def save_statistics_to_markdown(self, total_aig, missing_count, aig_percentage, td_percentage, match_percentage):
        """
        Save statistics to a markdown file as required by prompt.md
        """
        markdown_content = f"""# AIG Student Statistics Report

Generated on: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}

## Summary

- **Total students processed in Excel:** {self.stats['total_students']}
- **Students with AIG status:** {total_aig}
- **Students who are only TD:** {len(self.stats['td_only_students'])}

## AIG Status Breakdown

| Category | Count |
|----------|-------|
| Math only | {self.stats['aig_math_only']} |
| Reading only | {self.stats['aig_reading_only']} |
| Both Math & Reading | {self.stats['aig_both']} |
| No AIG status | {self.stats['aig_none']} |

## Source Document Analysis

- **Total students found in PDF/Word:** {len(self.students_in_sources)}
- **Students found in Excel:** {len(self.students_in_excel)}
- **Students in source docs but NOT in Excel:** {missing_count}

## Percentages

- **AIG students:** {aig_percentage:.1f}% of Excel total
- **TD-only students:** {td_percentage:.1f}% of Excel total
- **Students found in Excel:** {match_percentage:.1f}% of source total

## Generated Files

1. `updated_class_lists.xlsx` - Main Excel file with AIG columns and color coding
2. `updated_class_lists_AIG_Only.xlsx` - Excel file containing only AIG students
3. `students_not_in_excel.xlsx` - Students from PDF/Word not found in Excel

## Color Coding

- 🔵 **Light Blue (ADD8E6):** Students with AIG Math only
- 🟠 **Orange:** Students with AIG Reading only  
- 🟡 **Yellow:** Students with both AIG Math and Reading
"""
        
        # Save to output directory
        markdown_file = os.path.join(self.output_dir, 'aig_statistics_report.md')
        try:
            with open(markdown_file, 'w', encoding='utf-8') as f:
                f.write(markdown_content)
            logger.info(f"Statistics saved to markdown file: {markdown_file}")
        except Exception as e:
            logger.error(f"Error saving markdown file: {e}")

    def process(self):
        """
        Main processing method that orchestrates the entire workflow
        Updated for new format: updates existing Excel file with AIG columns
        """
        logger.info("Starting AIG Class List Processing...")
        
        try:
            # Step 1: Extract AIG students from PDF
            self.extract_aig_students_from_pdf()
            logger.info(f"Found {len(self.aig_students['math'])} AIG Math students")
            logger.info(f"Found {len(self.aig_students['reading'])} AIG Reading students")
            
            # Step 2: Extract AIG students from Word document (if provided)
            self.extract_aig_students_from_word()
            logger.info(f"Total {len(self.aig_students['math'])} AIG Math students after Word processing")
            logger.info(f"Total {len(self.aig_students['reading'])} AIG Reading students after Word processing")
            
            # Step 3: Update Excel file with AIG data
            self.update_excel_with_aig_data()
            
            # Step 4: Generate missing students report
            self.generate_missing_students_report()
            
            # Step 5: Print statistics
            self.print_statistics()
            
            logger.info("Processing completed successfully!")
            logger.info(f"Updated Excel file: {self.output_file}")
            
        except Exception as e:
            logger.error(f"Error during processing: {e}")
            raise


    def generate_missing_students_report(self):
        """
        Generate Excel report of students from PDF/Word who are not in Excel spreadsheet
        """
        logger.info("Generating missing students report...")
        
        # Find students in sources but not in Excel
        missing_students = self.students_in_sources - self.students_in_excel
        
        if not missing_students:
            logger.info("All students from source documents were found in Excel spreadsheet")
            return
        
        # Create workbook for missing students
        wb_missing = Workbook()
        
        # Remove default sheet and create new one
        if wb_missing.active:
            wb_missing.remove(wb_missing.active)
        
        ws_missing = wb_missing.create_sheet(title="Students Not In Excel")
        
        # Add headers
        headers = ['Student Name', 'Source', 'Student ID', 'Grade', 'AIG Math', 'AIG Reading', 'AIG Status']
        ws_missing.append(headers)
        
        # Add missing students data
        for student_name in sorted(missing_students):
            if hasattr(self, 'student_details') and student_name in self.student_details:
                student_info = self.student_details[student_name]
                
                # Determine source
                source = "PDF"
                if student_info['student_id'] == 'N/A':
                    source = "Word Document"
                
                # Determine AIG status
                if student_info['math'] and student_info['reading']:
                    aig_status = 'Both Math & Reading'
                    color = self.colors['both']
                elif student_info['math']:
                    aig_status = 'Math Only'
                    color = self.colors['math']
                elif student_info['reading']:
                    aig_status = 'Reading Only'
                    color = self.colors['reading']
                else:
                    aig_status = 'None'
                    color = None
                
                # Add row data
                row_data = [
                    student_name,
                    source,
                    student_info['student_id'],
                    student_info['grade'],
                    student_info['math'],
                    student_info['reading'],
                    aig_status
                ]
                
                ws_missing.append(row_data)
                
                # Apply color formatting
                if color:
                    row_num = ws_missing.max_row
                    for col_idx in range(1, 8):  # Columns A through G
                        ws_missing.cell(row=row_num, column=col_idx).fill = color
        
        # Auto-adjust column widths
        for column in ws_missing.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_missing.column_dimensions[column_letter].width = adjusted_width
        
        # Save the missing students report
        wb_missing.save(self.missing_students_file)
        logger.info(f"Missing students report saved: {self.missing_students_file}")
        logger.info(f"Found {len(missing_students)} students in source documents that are not in Excel")

def main():
    """
    Main function to run the application
    Updated to save files to output directory as required by prompt.md
    """
    # File paths - adjust these to match your files
    pdf_file = "input/SalemAIGRoster6.24.25.pdf"
    excel_file = "input/HEINZE of  25-26 Class Lists.xlsx"
    word_file = "input/TD from Finch WCPSS file.docx"  # Word document with additional AIG students
    output_dir = "output"  # Output directory as required by prompt.md
    
    # Check if files exist
    if not os.path.exists(pdf_file):
        logger.error(f"PDF file not found: {pdf_file}")
        return
    
    if not os.path.exists(excel_file):
        logger.error(f"Excel file not found: {excel_file}")
        return
    
    # Word file is optional
    if not os.path.exists(word_file):
        logger.warning(f"Word file not found: {word_file} - continuing with PDF only")
        word_file = None
    
    # Create processor and run
    processor = AIGClassListProcessor(pdf_file, excel_file, word_file, output_dir)
    processor.process()
    
    # Clean up temporary files
    temp_files = ["test_output.xlsx", "test_quick.py", "quick_test.py"]
    for temp_file in temp_files:
        if os.path.exists(temp_file):
            os.remove(temp_file)
            logger.info(f"Removed temporary file: {temp_file}")
    
    print("\n" + "="*50)
    print("AIG Class List Processing Complete!")
    print(f"Output directory: {output_dir}/")
    print(f"Main Excel file: {processor.output_file}")
    print(f"AIG-only Excel file: {processor.aig_only_file}")
    print(f"Missing students file: {processor.missing_students_file}")
    print("="*50)


if __name__ == "__main__":
    main()
